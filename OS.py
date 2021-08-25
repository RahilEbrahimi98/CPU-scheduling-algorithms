import openpyxl
import math
import time
import concurrent.futures
import matplotlib.pyplot
import tkinter
import numpy
from tkinter import filedialog

def read_from_file(name):
    global n
    global burst_time_one
    global burst_time_two
    global arrival_time 
    global io_time


    burst_time_one = []
    burst_time_two = []
    arrival_time = []
    global path
    io_time = []
    path = f"{name}"
    wb_obj = openpyxl.load_workbook(path) 
    sheet_obj = wb_obj.active
    n = sheet_obj.max_row - 1

    for i in range(2, n + 2):
        cell_obj = sheet_obj.cell(row = i, column = 1)
        temp_arrival_time = cell_obj.value
        arrival_time.append(temp_arrival_time)

    for i in range(2, n + 2):
        cell_obj = sheet_obj.cell(row = i, column = 2)
        temp_burst_time_one = cell_obj.value
        burst_time_one.append(temp_burst_time_one)
        
    for i in range(2, n + 2):
        cell_obj = sheet_obj.cell(row = i, column = 3)
        temp_io_time = cell_obj.value
        io_time.append(temp_io_time)

    for i in range(2, n + 2):
        cell_obj = sheet_obj.cell(row = i, column = 4)
        temp_burst_time_two = cell_obj.value
        burst_time_two.append(temp_burst_time_two)

    return

def write_to_file(name ,rt , ct , wt, tat):
    xfile = openpyxl.load_workbook(name)
    sheet = xfile.active
    sheet.cell(row=1, column=5).value = 'respones time'
    sheet.cell(row=1, column=6).value = 'completion time'
    sheet.cell(row=1, column=7).value = 'waiting time'
    sheet.cell(row=1, column=8).value = 'turn around time'
    
    for i in range(len(wt)):
        sheet.cell(row=i+2, column=5).value = rt[i]
        sheet.cell(row=i+2, column=6).value = ct[i]
        sheet.cell(row=i+2, column=7).value = wt[i]
        sheet.cell(row=i+2, column=8).value = tat[i]
        
    xfile.save(name)
    

def FCFS(arrival_time, burst_time_one, burst_time_two, io_time):
    process_timing = []
    current_time = 0
    ready_queue = []
    response_time = [0 for i in range(n)]
    completion_time = [0 for i in range(n)]
    flag = [0 for i in range(n)]
    for i in range(2 * n):
        min_index = arrival_time.index(min(arrival_time))
        if current_time <= arrival_time[min_index]:
            current_time = min(arrival_time)
        ready_queue.append(min_index)
        if flag[min_index] == 0:
            process_timing.append((ready_queue[-1], current_time, burst_time_one[min_index]))
            response_time[min_index] = current_time - arrival_time[min_index]
            arrival_time[min_index] = current_time + burst_time_one[min_index] + io_time[min_index]
            current_time += burst_time_one[min_index]
            flag[min_index] = 1
            if burst_time_two[min_index] == 0:
                completion_time[min_index] = current_time + io_time[min_index]
                flag[min_index] = 2
                arrival_time[min_index] = math.inf
        elif flag[min_index] == 1:
            process_timing.append((ready_queue[-1], current_time, burst_time_two[min_index]))
            arrival_time[min_index] = math.inf
            current_time += burst_time_two[min_index]
            flag[min_index] = 2
            completion_time[min_index] = current_time
    return [completion_time, response_time, process_timing]


def SJF(arrival_time, burst_time_one, burst_time_two, io_time):
    process_timing = []
    current_time = 0
    ready_queue = []
    response_time = [0 for i in range(n)]
    completion_time = [0 for i in range(n)]
    flag = [0 for i in range(n)]
    for i in range(2 * n):
        available_processes = [i for i in range(n) if arrival_time[i] <= current_time]
        if len(available_processes) == 0:
            current_time = min(arrival_time)
            available_processes = [i for i in range(n) if arrival_time[i] <= current_time]
        available_process_times = [math.inf for i in range(n)]
        for i in available_processes:
            if flag[i] == 0:
                available_process_times[i] = burst_time_one[i]
            elif flag[i] == 1:
                available_process_times[i] = burst_time_two[i]
        shortest_job = min(available_process_times)
        shortest_job_index = available_process_times.index(shortest_job)
        ready_queue.append(shortest_job_index)
        if flag[shortest_job_index] == 0:
            process_timing.append((ready_queue[-1], current_time, burst_time_one[shortest_job_index]))
            response_time[shortest_job_index] = current_time - arrival_time[shortest_job_index]
            arrival_time[shortest_job_index] = current_time + burst_time_one[shortest_job_index] + io_time[shortest_job_index]
            current_time += burst_time_one[shortest_job_index]
            flag[shortest_job_index] = 1
            if burst_time_two[shortest_job_index] == 0:
                completion_time[shortest_job_index] = current_time + io_time[shortest_job_index]
                flag[shortest_job_index] = 2
                arrival_time[shortest_job_index] = math.inf
        elif flag[shortest_job_index] == 1:
            process_timing.append((ready_queue[-1], current_time, burst_time_two[shortest_job_index]))
            arrival_time[shortest_job_index] = math.inf
            current_time += burst_time_two[shortest_job_index]
            flag[shortest_job_index] = 2
            completion_time[shortest_job_index] = current_time

    return [completion_time, response_time, process_timing]


def SRT(arrival_time, burst_time_one, burst_time_two, io_time):
    process_timing = []
    current_time = 0
    ready_queue = []
    response_time = [-1 for i in range(n)]
    completion_time = [0 for i in range(n)]
    flag = [0 for i in range(n)]
    end_flag = 0
    while end_flag == 0:
        available_processes = [i for i in range(n) if arrival_time[i] <= current_time]
        if len(available_processes) == 0:
            current_time = min(arrival_time)
            available_processes = [i for i in range(n) if arrival_time[i] <= current_time]
        available_process_times = [math.inf for i in range(n)]
        for i in available_processes:
            if flag[i] == 0:
                available_process_times[i] = burst_time_one[i]
            elif flag[i] == 1:
                available_process_times[i] = burst_time_two[i]
        shortest_job = min(available_process_times)
        shortest_job_index = available_process_times.index(shortest_job)
        ready_queue.append(shortest_job_index)
        if response_time[shortest_job_index] == -1:
            response_time[shortest_job_index] = current_time - arrival_time[shortest_job_index]
        new_arrivals = [i for i in arrival_time if i < current_time+shortest_job and i > current_time]
        if len(new_arrivals) > 0:
            new_arrivals_index = [arrival_time.index(i) for i in new_arrivals]
            new_burst_times = [0 for i in range(len(new_arrivals))]
            for i in range(len(new_arrivals_index)):
                if flag[new_arrivals_index[i]]==0:
                    new_burst_times[i] = burst_time_one[new_arrivals_index[i]]
                elif flag[new_arrivals_index[i]]==1:
                    new_burst_times[i] = burst_time_two[new_arrivals_index[i]]
            new_available_index=[]
            new_available_arrival=[]
            new_available_burst_time=[]
            for i in range(len(new_arrivals)):
                if shortest_job + current_time - new_arrivals[i] > new_burst_times[i]:
                    new_available_index.append(new_arrivals_index[i])
                    new_available_arrival.append(new_arrivals[i])
                    new_available_burst_time.append(new_burst_times[i])
            if len(new_available_index) > 0:
                new_shortest_job_arrival_time = min(new_arrivals)
                new_min_arrival_time_indexes = [new_available_index[i] for i in range(len(new_available_arrival)) if new_available_arrival[i] == new_shortest_job_arrival_time]
                new_min_arrival_time_burst_times = [new_available_burst_time[i] for i in range(len(new_available_arrival)) if new_available_arrival[i] == new_shortest_job_arrival_time]
                min_burst_time = min(new_min_arrival_time_burst_times)
                min_burst_time_index = new_min_arrival_time_indexes[new_min_arrival_time_burst_times.index(min_burst_time)]
                process_timing.append((ready_queue[-1], current_time, new_shortest_job_arrival_time - current_time))
                current_time = new_shortest_job_arrival_time
                if flag[shortest_job_index] == 0:
                    burst_time_one[shortest_job_index] -= current_time
                elif flag[shortest_job_index] == 1:
                    burst_time_two[shortest_job_index] -= current_time
                continue
        if flag[shortest_job_index] == 0:
            process_timing.append((ready_queue[-1], current_time, burst_time_one[shortest_job_index]))
            arrival_time[shortest_job_index] = current_time + burst_time_one[shortest_job_index] + io_time[shortest_job_index]
            current_time += burst_time_one[shortest_job_index]
            flag[shortest_job_index] = 1
            if burst_time_two[shortest_job_index] == 0:
                completion_time[shortest_job_index] = current_time + io_time[shortest_job_index]
                arrival_time[shortest_job_index] = math.inf
                flag[shortest_job_index] = 2
        elif flag[shortest_job_index] == 1:
            process_timing.append((ready_queue[-1], current_time, burst_time_two[shortest_job_index]))
            arrival_time[shortest_job_index] = math.inf
            current_time += burst_time_two[shortest_job_index]
            flag[shortest_job_index] = 2
            completion_time[shortest_job_index] = current_time
        end_flag = 1
        for time in arrival_time:
            if time != math.inf:
                end_flag = 0
    return [completion_time, response_time, process_timing]


def RR(arrival_time, burst_time_one, burst_time_two, io_time, time_quantum):
    process_timing = []
    current_time = 0
    ready_queue = []
    available_processes = []
    response_time = [-1 for i in range(n)]
    completion_time = [0 for i in range(n)]
    flag = [0 for i in range(n)]
    end_flag = 0
    current_time = min(arrival_time)
    available_processes = [i for i in range(len(arrival_time)) if arrival_time[i] == current_time]
    while end_flag == 0:
        process_flag = 0
        current_process = available_processes.pop(0)
        if response_time[current_process] == -1:
            response_time[current_process] = current_time - arrival_time[current_process]
        ready_queue.append(current_process)
        if flag[current_process] == 0:
            if burst_time_one[current_process] <= time_quantum:
                process_timing.append((ready_queue[-1], current_time, burst_time_one[current_process]))
                current_time += burst_time_one[current_process]
                flag[current_process] = 1
                arrival_time[current_process] = current_time + io_time[current_process]
                process_flag = 1
                if burst_time_two[current_process] == 0:
                    completion_time[current_process] = current_time + io_time[current_process]
                    flag[current_process] = 2
                    arrival_time[current_process] = math.inf
            elif burst_time_one[current_process] > time_quantum:
                process_timing.append((ready_queue[-1], current_time, time_quantum))
                current_time += time_quantum
                burst_time_one[current_process] -= time_quantum
                arrival_time[current_process] = current_time
        elif flag[current_process] == 1:
            if burst_time_two[current_process] <= time_quantum:
                process_timing.append((ready_queue[-1], current_time, burst_time_two[current_process]))
                current_time += burst_time_two[current_process]
                completion_time[current_process] = current_time
                flag[current_process] = 2
                arrival_time[current_process] = math.inf
                process_flag = 1
            elif burst_time_two[current_process] > time_quantum:
                process_timing.append((ready_queue[-1], current_time, time_quantum))
                current_time += time_quantum
                burst_time_two[current_process] -= time_quantum
                arrival_time[current_process] = current_time
        new_processes = [(process, arrival_time[process]) for process in range(len(arrival_time)) if (arrival_time[process] > current_time - time_quantum and arrival_time[process] < current_time)]
        new_processes.sort(key=lambda tup: tup[1])
        for new_process in new_processes:
            if new_process not in available_processes:
                available_processes.append(new_process[0])
        for i in range(len(arrival_time)):
            if arrival_time[i] == current_time:
                available_processes.append(i)
        if current_process in available_processes:
            available_processes.pop(available_processes.index(current_process))
        if process_flag == 0:
            available_processes.append(current_process)
        for process in available_processes:
            if arrival_time[process] == math.inf:
                available_processes.pop(available_processes.index(process))
        if len(available_processes) == 0:
            current_time = min(arrival_time)
            available_processes = [i for i in range(len(arrival_time)) if arrival_time[i] == current_time]
        end_flag = 1
        for time in arrival_time:
            if time != math.inf:
                end_flag = 0
    return [completion_time, response_time, process_timing]

def waiting_time(arrival_time, burst_time_one, burst_time_two, io_time, completion_time):
    waiting_time = [0 for i in range(n)]
    for i in range(n):
        waiting_time[i] = completion_time[i] - (arrival_time[i] + burst_time_one[i] + burst_time_two[i] + io_time[i])
    return waiting_time


def turnaround(completion_time, arrival_time):
    turnaround_time = [0 for i in range(n)]
    for i in range(n):
        turnaround_time[i] = completion_time[i] - arrival_time[i]
    return turnaround_time


def utilization(completion_time, burst_time_one, burst_time_two):
    return (sum(burst_time_one) + sum(burst_time_two)) / max(completion_time)


def throughput(completion_time):
    return n / max(completion_time)

def RR_all(time_quantum , flag):
    RR_completion_response_time = RR([i for i in arrival_time], [i for i in burst_time_one], [i for i in burst_time_two], io_time, time_quantum)
    RR_completion_time          = RR_completion_response_time[0]
    RR_response_time            = RR_completion_response_time[1]
    RR_process_time             = RR_completion_response_time[2]
    RR_waiting_time             = waiting_time(arrival_time, burst_time_one, burst_time_two, io_time, RR_completion_time)
    RR_turnaround_time          = turnaround(RR_completion_time, arrival_time)
    RR_utilization              = utilization(RR_completion_time, burst_time_one, burst_time_two)
    RR_throughput               = throughput(RR_completion_time)
    
    if flag==0:
        write_to_file(path , RR_response_time , RR_completion_time ,RR_waiting_time , RR_turnaround_time )
    
    return {'process_time' : RR_process_time,'completion_time': RR_completion_time, 'response_time': RR_response_time, 'waiting_time': RR_waiting_time, 'turnaround_time': RR_turnaround_time, 'utilization': RR_utilization, 'throughput': RR_throughput}

def SRT_all(flag):
    SRT_completion_response_time = SRT([i for i in arrival_time], [i for i in burst_time_one], [i for i in burst_time_two], io_time)
    SRT_completion_time          = SRT_completion_response_time[0]
    SRT_response_time            = SRT_completion_response_time[1]
    SRT_process_time             = SRT_completion_response_time[2]
    SRT_waiting_time             = waiting_time(arrival_time, burst_time_one, burst_time_two, io_time, SRT_completion_time)
    SRT_turnaround_time          = turnaround(SRT_completion_time, arrival_time)
    SRT_utilization              = utilization(SRT_completion_time, burst_time_one, burst_time_two)
    SRT_throughput               = throughput(SRT_completion_time)
    
    if flag==0:
        write_to_file(path , SRT_response_time , SRT_completion_time ,SRT_waiting_time , SRT_turnaround_time )
    
    return {'process_time' : SRT_process_time,'completion_time': SRT_completion_time, 'response_time': SRT_response_time, 'waiting_time': SRT_waiting_time, 'turnaround_time': SRT_turnaround_time, 'utilization': SRT_utilization, 'throughput': SRT_throughput}

def SJF_all(flag):
    SJF_completion_response_time = SJF([i for i in arrival_time], [i for i in burst_time_one], [i for i in burst_time_two], io_time)
    SJF_completion_time          = SJF_completion_response_time[0]
    SJF_response_time            = SJF_completion_response_time[1]
    SJF_process_time             = SJF_completion_response_time[2]
    SJF_waiting_time             = waiting_time(arrival_time, burst_time_one, burst_time_two, io_time, SJF_completion_time)
    SJF_turnaround_time          = turnaround(SJF_completion_time, arrival_time)
    SJF_utilization              = utilization(SJF_completion_time, burst_time_one, burst_time_two)
    SJF_throughput               = throughput(SJF_completion_time)
    
    if flag==0:
        write_to_file(path , SJF_response_time , SJF_completion_time ,SJF_waiting_time , SJF_turnaround_time )
    
    return {'process_time' : SJF_process_time,'completion_time': SJF_completion_time, 'response_time': SJF_response_time, 'waiting_time': SJF_waiting_time, 'turnaround_time': SJF_turnaround_time, 'utilization': SJF_utilization, 'throughput': SJF_throughput}

def FCFS_all(flag):
    FCFS_completion_response_time = FCFS([i for i in arrival_time], [i for i in burst_time_one], [i for i in burst_time_two], io_time)
    FCFS_completion_time          = FCFS_completion_response_time[0]
    FCFS_response_time            = FCFS_completion_response_time[1]
    FCFS_process_time             = FCFS_completion_response_time[2]
    FCFS_waiting_time             = waiting_time(arrival_time, burst_time_one, burst_time_two, io_time, FCFS_completion_time)
    FCFS_turnaround_time          = turnaround(FCFS_completion_time, arrival_time)
    FCFS_utilization              = utilization(FCFS_completion_time, burst_time_one, burst_time_two)
    FCFS_throughput               = throughput(FCFS_completion_time)
    
    if flag==0:
        write_to_file(path , FCFS_response_time , FCFS_completion_time ,FCFS_waiting_time , FCFS_turnaround_time )
    
    return {'process_time' : FCFS_process_time, 'completion_time': FCFS_completion_time, 'response_time': FCFS_response_time, 'waiting_time': FCFS_waiting_time, 'turnaround_time': FCFS_turnaround_time, 'utilization': FCFS_utilization, 'throughput': FCFS_throughput}

def RR_time(time_quantum ,flag):
    time_start = time.perf_counter()
    RR_all(time_quantum , flag)
    time_end = time.perf_counter()
    return round(time_end - time_start, 5)

def SRT_time(flag):
    time_start = time.perf_counter()
    SRT_all(flag)
    time_end = time.perf_counter()
    return round(time_end - time_start, 5)

def SJF_time(flag):
    time_start = time.perf_counter()
    SJF_all(flag)
    time_end = time.perf_counter()
    return round(time_end - time_start, 5)

def FCFS_time(flag):
    time_start = time.perf_counter()
    FCFS_all(flag)
    time_end = time.perf_counter()
    
    return round(time_end - time_start, 5)

def run_all(time_quantum):
    with concurrent.futures.ThreadPoolExecutor() as executor:
        FCFS_result_object = executor.submit(FCFS_all , 1)
        SJF_result_object = executor.submit(SJF_all , 1)
        SRT_result_object = executor.submit(SRT_all , 1)
        RR_result_object = executor.submit(RR_all, time_quantum ,1)
        
        results = {'FCFS': FCFS_result_object.result(), 'SJF': SJF_result_object.result(), 'SRT': SRT_result_object.result(), 'RR': RR_result_object.result()}
    return results

def run_all_time(time_quantum):
    time_start = time.perf_counter()
    run_all(time_quantum)
    time_end = time.perf_counter()
    return round(time_end - time_start, 5)

def draw_gantt(process_timing):
    fig, gnt = matplotlib.pyplot.subplots() 
    gnt.set_ylim(0, n * 10) 
    gnt.set_xlim(-2, process_timing[-1][1] + process_timing[-1][2] + 2) 
    gnt.set_xlabel('seconds since start') 
    gnt.set_ylabel('Process') 
    gnt.set_yticks([n * 5]) 
    gnt.set_yticklabels(['Timeline']) 
    gnt.grid(True)
    gnt.broken_barh([(process_timing[i][1], process_timing[i][2]) for i in range(len(process_timing)) if process_timing[i][0] == 0], (3.75 * n, 10), facecolors =('#ff9ff3'), label='P0')
    gnt.broken_barh([(process_timing[i][1], process_timing[i][2]) for i in range(len(process_timing)) if process_timing[i][0] == 1], (3.75 * n, 10), facecolors =('#feca57'), label='P1') 
    gnt.broken_barh([(process_timing[i][1], process_timing[i][2]) for i in range(len(process_timing)) if process_timing[i][0] == 2], (3.75 * n, 10), facecolors =('#ff6b6b'), label='P2') 
    gnt.broken_barh([(process_timing[i][1], process_timing[i][2]) for i in range(len(process_timing)) if process_timing[i][0] == 3], (3.75 * n, 10), facecolors =('#48dbfb'), label='P3') 
    gnt.broken_barh([(process_timing[i][1], process_timing[i][2]) for i in range(len(process_timing)) if process_timing[i][0] == 4], (3.75 * n, 10), facecolors =('#1dd1a1'), label='P4') 
    gnt.broken_barh([(process_timing[i][1], process_timing[i][2]) for i in range(len(process_timing)) if process_timing[i][0] == 5], (3.75 * n, 10), facecolors =('#00d2d3'), label='P5') 
    gnt.broken_barh([(process_timing[i][1], process_timing[i][2]) for i in range(len(process_timing)) if process_timing[i][0] == 6], (3.75 * n, 10), facecolors =('#54a0ff'), label='P6') 
    gnt.broken_barh([(process_timing[i][1], process_timing[i][2]) for i in range(len(process_timing)) if process_timing[i][0] == 7], (3.75 * n, 10), facecolors =('#5f27cd'), label='P7') 
    gnt.broken_barh([(process_timing[i][1], process_timing[i][2]) for i in range(len(process_timing)) if process_timing[i][0] == 8], (3.75 * n, 10), facecolors =('#c8d6e5'), label='P8') 
    gnt.broken_barh([(process_timing[i][1], process_timing[i][2]) for i in range(len(process_timing)) if process_timing[i][0] == 9], (3.75 * n, 10), facecolors =('#576574'), label='P9')
    mng = matplotlib.pyplot.get_current_fig_manager()
    mng.resize(*mng.window.maxsize())
    matplotlib.pyplot.title('Gantt Chart')
    gnt.legend(loc='best', shadow=True, ncol=2, prop={'size': 16})
    matplotlib.pyplot.show() 
def draw_all_gantt(FCFS_process_timing , SJF_process_timing , SRT_process_timing , RR_process_timing):
    fig, gnt  = matplotlib.pyplot.subplots() 
    gnt.set_ylim(0, 31) 
    gnt.set_xlim(-2, FCFS_process_timing[-1][1] + FCFS_process_timing[-1][2] + 2) 
    gnt.set_xlabel('seconds since start') 
    gnt.set_ylabel('Process') 
    gnt.set_yticks([3.5, 11.5 , 19.5 , 27.5]) 
    gnt.set_yticklabels(['FCFS Timeline' , 'SJF Timeline' , 'SRT Timeline' , 'RR Timeline']) 
    gnt.grid(True)
    gnt.broken_barh([(FCFS_process_timing[i][1], FCFS_process_timing[i][2]) for i in range(len(FCFS_process_timing)) if FCFS_process_timing[i][0] == 0], (0, 7), facecolors =('#ff9ff3'), label='P0')
    gnt.broken_barh([(FCFS_process_timing[i][1], FCFS_process_timing[i][2]) for i in range(len(FCFS_process_timing)) if FCFS_process_timing[i][0] == 1], (0, 7), facecolors =('#feca57'), label='P1') 
    gnt.broken_barh([(FCFS_process_timing[i][1], FCFS_process_timing[i][2]) for i in range(len(FCFS_process_timing)) if FCFS_process_timing[i][0] == 2], (0, 7), facecolors =('#ff6b6b'), label='P2') 
    gnt.broken_barh([(FCFS_process_timing[i][1], FCFS_process_timing[i][2]) for i in range(len(FCFS_process_timing)) if FCFS_process_timing[i][0] == 3], (0, 7), facecolors =('#48dbfb'), label='P3') 
    gnt.broken_barh([(FCFS_process_timing[i][1], FCFS_process_timing[i][2]) for i in range(len(FCFS_process_timing)) if FCFS_process_timing[i][0] == 4], (0, 7), facecolors =('#1dd1a1'), label='P4') 
    gnt.broken_barh([(FCFS_process_timing[i][1], FCFS_process_timing[i][2]) for i in range(len(FCFS_process_timing)) if FCFS_process_timing[i][0] == 5], (0, 7), facecolors =('#00d2d3'), label='P5') 
    gnt.broken_barh([(FCFS_process_timing[i][1], FCFS_process_timing[i][2]) for i in range(len(FCFS_process_timing)) if FCFS_process_timing[i][0] == 6], (0, 7), facecolors =('#54a0ff'), label='P6') 
    gnt.broken_barh([(FCFS_process_timing[i][1], FCFS_process_timing[i][2]) for i in range(len(FCFS_process_timing)) if FCFS_process_timing[i][0] == 7], (0, 7), facecolors =('#5f27cd'), label='P7') 
    gnt.broken_barh([(FCFS_process_timing[i][1], FCFS_process_timing[i][2]) for i in range(len(FCFS_process_timing)) if FCFS_process_timing[i][0] == 8], (0, 7), facecolors =('#c8d6e5'), label='P8') 
    gnt.broken_barh([(FCFS_process_timing[i][1], FCFS_process_timing[i][2]) for i in range(len(FCFS_process_timing)) if FCFS_process_timing[i][0] == 9], (0, 7), facecolors =('#576574'), label='P9')
    
    gnt.broken_barh([(SJF_process_timing[i][1], SJF_process_timing[i][2]) for i in range(len(SJF_process_timing)) if SJF_process_timing[i][0] == 0], (8, 7), facecolors =('#ff9ff3'))
    gnt.broken_barh([(SJF_process_timing[i][1], SJF_process_timing[i][2]) for i in range(len(SJF_process_timing)) if SJF_process_timing[i][0] == 1], (8, 7), facecolors =('#feca57')) 
    gnt.broken_barh([(SJF_process_timing[i][1], SJF_process_timing[i][2]) for i in range(len(SJF_process_timing)) if SJF_process_timing[i][0] == 2], (8, 7), facecolors =('#ff6b6b')) 
    gnt.broken_barh([(SJF_process_timing[i][1], SJF_process_timing[i][2]) for i in range(len(SJF_process_timing)) if SJF_process_timing[i][0] == 3], (8, 7), facecolors =('#48dbfb')) 
    gnt.broken_barh([(SJF_process_timing[i][1], SJF_process_timing[i][2]) for i in range(len(SJF_process_timing)) if SJF_process_timing[i][0] == 4], (8, 7), facecolors =('#1dd1a1')) 
    gnt.broken_barh([(SJF_process_timing[i][1], SJF_process_timing[i][2]) for i in range(len(SJF_process_timing)) if SJF_process_timing[i][0] == 5], (8, 7), facecolors =('#00d2d3')) 
    gnt.broken_barh([(SJF_process_timing[i][1], SJF_process_timing[i][2]) for i in range(len(SJF_process_timing)) if SJF_process_timing[i][0] == 6], (8, 7), facecolors =('#54a0ff')) 
    gnt.broken_barh([(SJF_process_timing[i][1], SJF_process_timing[i][2]) for i in range(len(SJF_process_timing)) if SJF_process_timing[i][0] == 7], (8, 7), facecolors =('#5f27cd')) 
    gnt.broken_barh([(SJF_process_timing[i][1], SJF_process_timing[i][2]) for i in range(len(SJF_process_timing)) if SJF_process_timing[i][0] == 8], (8, 7), facecolors =('#c8d6e5')) 
    gnt.broken_barh([(SJF_process_timing[i][1], SJF_process_timing[i][2]) for i in range(len(SJF_process_timing)) if SJF_process_timing[i][0] == 9], (8, 7), facecolors =('#576574'))
    
    gnt.broken_barh([(SRT_process_timing[i][1], SRT_process_timing[i][2]) for i in range(len(SRT_process_timing)) if SRT_process_timing[i][0] == 0], (16, 7), facecolors =('#ff9ff3'))
    gnt.broken_barh([(SRT_process_timing[i][1], SRT_process_timing[i][2]) for i in range(len(SRT_process_timing)) if SRT_process_timing[i][0] == 1], (16, 7), facecolors =('#feca57')) 
    gnt.broken_barh([(SRT_process_timing[i][1], SRT_process_timing[i][2]) for i in range(len(SRT_process_timing)) if SRT_process_timing[i][0] == 2], (16, 7), facecolors =('#ff6b6b')) 
    gnt.broken_barh([(SRT_process_timing[i][1], SRT_process_timing[i][2]) for i in range(len(SRT_process_timing)) if SRT_process_timing[i][0] == 3], (16, 7), facecolors =('#48dbfb')) 
    gnt.broken_barh([(SRT_process_timing[i][1], SRT_process_timing[i][2]) for i in range(len(SRT_process_timing)) if SRT_process_timing[i][0] == 4], (16, 7), facecolors =('#1dd1a1')) 
    gnt.broken_barh([(SRT_process_timing[i][1], SRT_process_timing[i][2]) for i in range(len(SRT_process_timing)) if SRT_process_timing[i][0] == 5], (16, 7), facecolors =('#00d2d3')) 
    gnt.broken_barh([(SRT_process_timing[i][1], SRT_process_timing[i][2]) for i in range(len(SRT_process_timing)) if SRT_process_timing[i][0] == 6], (16, 7), facecolors =('#54a0ff')) 
    gnt.broken_barh([(SRT_process_timing[i][1], SRT_process_timing[i][2]) for i in range(len(SRT_process_timing)) if SRT_process_timing[i][0] == 7], (16, 7), facecolors =('#5f27cd')) 
    gnt.broken_barh([(SRT_process_timing[i][1], SRT_process_timing[i][2]) for i in range(len(SRT_process_timing)) if SRT_process_timing[i][0] == 8], (16, 7), facecolors =('#c8d6e5')) 
    gnt.broken_barh([(SRT_process_timing[i][1], SRT_process_timing[i][2]) for i in range(len(SRT_process_timing)) if SRT_process_timing[i][0] == 9], (16, 7), facecolors =('#576574'))
    
    gnt.broken_barh([(RR_process_timing[i][1], RR_process_timing[i][2]) for i in range(len(RR_process_timing)) if RR_process_timing[i][0] == 0], (24, 7), facecolors =('#ff9ff3'))
    gnt.broken_barh([(RR_process_timing[i][1], RR_process_timing[i][2]) for i in range(len(RR_process_timing)) if RR_process_timing[i][0] == 1], (24, 7), facecolors =('#feca57')) 
    gnt.broken_barh([(RR_process_timing[i][1], RR_process_timing[i][2]) for i in range(len(RR_process_timing)) if RR_process_timing[i][0] == 2], (24, 7), facecolors =('#ff6b6b')) 
    gnt.broken_barh([(RR_process_timing[i][1], RR_process_timing[i][2]) for i in range(len(RR_process_timing)) if RR_process_timing[i][0] == 3], (24, 7), facecolors =('#48dbfb')) 
    gnt.broken_barh([(RR_process_timing[i][1], RR_process_timing[i][2]) for i in range(len(RR_process_timing)) if RR_process_timing[i][0] == 4], (24, 7), facecolors =('#1dd1a1')) 
    gnt.broken_barh([(RR_process_timing[i][1], RR_process_timing[i][2]) for i in range(len(RR_process_timing)) if RR_process_timing[i][0] == 5], (24, 7), facecolors =('#00d2d3')) 
    gnt.broken_barh([(RR_process_timing[i][1], RR_process_timing[i][2]) for i in range(len(RR_process_timing)) if RR_process_timing[i][0] == 6], (24, 7), facecolors =('#54a0ff')) 
    gnt.broken_barh([(RR_process_timing[i][1], RR_process_timing[i][2]) for i in range(len(RR_process_timing)) if RR_process_timing[i][0] == 7], (24, 7), facecolors =('#5f27cd')) 
    gnt.broken_barh([(RR_process_timing[i][1], RR_process_timing[i][2]) for i in range(len(RR_process_timing)) if RR_process_timing[i][0] == 8], (24, 7), facecolors =('#c8d6e5')) 
    gnt.broken_barh([(RR_process_timing[i][1], RR_process_timing[i][2]) for i in range(len(RR_process_timing)) if RR_process_timing[i][0] == 9], (24, 7), facecolors =('#576574'))
    
    mng = matplotlib.pyplot.get_current_fig_manager()
    mng.resize(*mng.window.maxsize())
    matplotlib.pyplot.title('all Gantt Chart')
    gnt.legend(loc='best', shadow=True, ncol=1, prop={'size': 9})
    matplotlib.pyplot.show() 
    
def draw_FCFS_chart():
    timing = FCFS([i for i in arrival_time], [i for i in burst_time_one], [i for i in burst_time_two], io_time)
    process_timing = timing[2]
    draw_gantt(process_timing)
    return

def draw_SJF_chart():
    timing = SJF([i for i in arrival_time], [i for i in burst_time_one], [i for i in burst_time_two], io_time)
    process_timing = timing[2]
    draw_gantt(process_timing)
    return

def draw_SRT_chart():
    timing = SRT([i for i in arrival_time], [i for i in burst_time_one], [i for i in burst_time_two], io_time)
    process_timing = timing[2]
    draw_gantt(process_timing)
    return

def draw_RR_chart(time_quantum):
    timing = RR([i for i in arrival_time], [i for i in burst_time_one], [i for i in burst_time_two], io_time, time_quantum)
    process_timing = timing[2]
    draw_gantt(process_timing)
    return

def draw_All_algorithms_chart(time_quantum):
    dictionary = run_all(time_quantum)
    FCFS=dictionary['FCFS']
    SJF=dictionary['SJF']
    SRT=dictionary['SRT']
    RR=dictionary['RR']
    draw_all_gantt(FCFS['process_time'] , SJF['process_time'], SRT['process_time'] , RR['process_time'])
    return 
    

def FCFS_process_data_chart(flag):
    data_for_chart = FCFS_all(flag)
    completion_time = data_for_chart['completion_time']
    response_time = data_for_chart['response_time']
    waiting_time = data_for_chart['waiting_time']
    turnaround_time = data_for_chart['turnaround_time']
    matplotlib.pyplot.grid(True)
    matplotlib.pyplot.subplot(2, 2, 1) 
    matplotlib.pyplot.title('Completion Time') 
    matplotlib.pyplot.xlabel('Process Number') 
    matplotlib.pyplot.ylabel('Time') 
    matplotlib.pyplot.xticks([i for i in range(n)], [f'P{i}' for i in range(n)])
    for i, v in enumerate(completion_time):
        matplotlib.pyplot.text(i - 0.02, v + max(completion_time) * 0.01, f'{v}', color='blue', fontweight='bold')
    matplotlib.pyplot.grid(True)
    matplotlib.pyplot.bar([i for i in range(n)], completion_time, 0.5)
    matplotlib.pyplot.subplot(2, 2, 2) 
    matplotlib.pyplot.title('Response Time') 
    matplotlib.pyplot.xlabel('Process Number') 
    matplotlib.pyplot.ylabel('Time') 
    matplotlib.pyplot.xticks([i for i in range(n)], [f'P{i}' for i in range(n)])
    for i, v in enumerate(response_time):
        matplotlib.pyplot.text(i - 0.02, v + max(response_time) * 0.01, f'{v}', color='blue', fontweight='bold')
    matplotlib.pyplot.grid(True)
    matplotlib.pyplot.bar([i for i in range(n)], response_time, 0.5)
    matplotlib.pyplot.subplot(2, 2, 3) 
    matplotlib.pyplot.title('Waiting Time') 
    matplotlib.pyplot.xlabel('Process Number') 
    matplotlib.pyplot.ylabel('Time') 
    matplotlib.pyplot.xticks([i for i in range(n)], [f'P{i}' for i in range(n)])
    for i, v in enumerate(waiting_time):
        matplotlib.pyplot.text(i - 0.02, v + max(waiting_time) * 0.01, f'{v}', color='blue', fontweight='bold')
    matplotlib.pyplot.grid(True)
    matplotlib.pyplot.bar([i for i in range(n)], waiting_time, 0.5)
    matplotlib.pyplot.subplot(2, 2, 4) 
    matplotlib.pyplot.title('Turnaround Time') 
    matplotlib.pyplot.xlabel('Process Number') 
    matplotlib.pyplot.ylabel('Time') 
    matplotlib.pyplot.xticks([i for i in range(n)], [f'P{i}' for i in range(n)])
    for i, v in enumerate(turnaround_time):
        matplotlib.pyplot.text(i - 0.02, v + max(turnaround_time) * 0.01, f'{v}', color='blue', fontweight='bold')
    matplotlib.pyplot.bar([i for i in range(n)], turnaround_time, 0.5)
    mng = matplotlib.pyplot.get_current_fig_manager()
    mng.resize(*mng.window.maxsize())
    matplotlib.pyplot.grid(True)
    matplotlib.pyplot.show()

def FCFS_algorithm_data_chart(flag):
    data_for_chart = FCFS_all(flag)
    completion_time = data_for_chart['completion_time']
    response_time = data_for_chart['response_time']
    waiting_time = data_for_chart['waiting_time']
    turnaround_time = data_for_chart['turnaround_time']
    utilization = data_for_chart['utilization']
    throughput = data_for_chart['throughput']
    run_time = FCFS_time(flag) * 1000
    matplotlib.pyplot.title('Algorithm Data') 
    matplotlib.pyplot.xlabel('Name') 
    matplotlib.pyplot.ylabel('Number') 
    matplotlib.pyplot.xticks([i for i in range(6)], ['Utilization', 'Throughput', 'Avg. Response Time', 'Avg. Waiting Time', 'Avg. Turnaround Time', 'Run Time * 1000'])
    for i, v in enumerate([utilization, throughput, sum(response_time) / len(response_time), sum(waiting_time) / len(waiting_time), sum(turnaround_time) / len(turnaround_time), run_time]):
        matplotlib.pyplot.text(i - 0.02, v + max([utilization, throughput, sum(response_time) / len(response_time), sum(waiting_time) / len(waiting_time), sum(turnaround_time) / len(turnaround_time), run_time]) * 0.01, f'{round(v, 2)}', color='blue', fontweight='bold')
    matplotlib.pyplot.grid(True)
    matplotlib.pyplot.bar([i for i in range(6)], [utilization, throughput, sum(response_time) / len(response_time), sum(waiting_time) / len(waiting_time), sum(turnaround_time) / len(turnaround_time), run_time], 0.5)
    mng = matplotlib.pyplot.get_current_fig_manager()
    mng.resize(*mng.window.maxsize())
    matplotlib.pyplot.show()
    
def SJF_process_data_chart(flag):
    data_for_chart = SJF_all(flag)
    completion_time = data_for_chart['completion_time']
    response_time = data_for_chart['response_time']
    waiting_time = data_for_chart['waiting_time']
    turnaround_time = data_for_chart['turnaround_time']
    matplotlib.pyplot.subplot(2, 2, 1) 
    matplotlib.pyplot.title('Completion Time') 
    matplotlib.pyplot.xlabel('Process Number') 
    matplotlib.pyplot.ylabel('Time') 
    matplotlib.pyplot.xticks([i for i in range(n)], [f'P{i}' for i in range(n)])
    matplotlib.pyplot.grid(True)
    for i, v in enumerate(completion_time):
        matplotlib.pyplot.text(i - 0.02, v + max(completion_time) * 0.01, f'{v}', color='blue', fontweight='bold')
    matplotlib.pyplot.bar([i for i in range(n)], completion_time, 0.5)
    matplotlib.pyplot.subplot(2, 2, 2) 
    matplotlib.pyplot.title('Response Time') 
    matplotlib.pyplot.xlabel('Process Number') 
    matplotlib.pyplot.ylabel('Time') 
    matplotlib.pyplot.xticks([i for i in range(n)], [f'P{i}' for i in range(n)])
    matplotlib.pyplot.grid(True)
    for i, v in enumerate(response_time):
        matplotlib.pyplot.text(i - 0.02, v + max(response_time) * 0.01, f'{v}', color='blue', fontweight='bold')
    matplotlib.pyplot.bar([i for i in range(n)], response_time, 0.5)
    matplotlib.pyplot.subplot(2, 2, 3) 
    matplotlib.pyplot.title('Waiting Time') 
    matplotlib.pyplot.xlabel('Process Number') 
    matplotlib.pyplot.ylabel('Time') 
    matplotlib.pyplot.xticks([i for i in range(n)], [f'P{i}' for i in range(n)])
    matplotlib.pyplot.grid(True)
    for i, v in enumerate(waiting_time):
        matplotlib.pyplot.text(i - 0.02, v + max(waiting_time) * 0.01, f'{v}', color='blue', fontweight='bold')
    matplotlib.pyplot.bar([i for i in range(n)], waiting_time, 0.5)
    matplotlib.pyplot.subplot(2, 2, 4) 
    matplotlib.pyplot.title('Turnaround Time') 
    matplotlib.pyplot.xlabel('Process Number') 
    matplotlib.pyplot.ylabel('Time') 
    matplotlib.pyplot.xticks([i for i in range(n)], [f'P{i}' for i in range(n)])
    matplotlib.pyplot.grid(True)
    for i, v in enumerate(turnaround_time):
        matplotlib.pyplot.text(i - 0.02, v + max(turnaround_time) * 0.01, f'{v}', color='blue', fontweight='bold')
    matplotlib.pyplot.bar([i for i in range(n)], turnaround_time, 0.5)
    mng = matplotlib.pyplot.get_current_fig_manager()
    mng.resize(*mng.window.maxsize())
    matplotlib.pyplot.show()

def SJF_algorithm_data_chart(flag):
    data_for_chart = SJF_all(flag)
    utilization = data_for_chart['utilization']
    throughput = data_for_chart['throughput']
    completion_time = data_for_chart['completion_time']
    response_time = data_for_chart['response_time']
    waiting_time = data_for_chart['waiting_time']
    turnaround_time = data_for_chart['turnaround_time']
    run_time = SJF_time(flag) * 1000
    matplotlib.pyplot.title('Algorithm Data') 
    matplotlib.pyplot.xlabel('Name') 
    matplotlib.pyplot.ylabel('Number') 
    matplotlib.pyplot.xticks([i for i in range(6)], ['Utilization', 'Throughput', 'Avg. Response Time', 'Avg. Waiting Time', 'Avg. Turnaround Time', 'Run Time * 1000'])
    matplotlib.pyplot.grid(True)
    for i, v in enumerate([utilization, throughput, sum(response_time) / len(response_time), sum(waiting_time) / len(waiting_time), sum(turnaround_time) / len(turnaround_time), run_time]):
        matplotlib.pyplot.text(i - 0.02, v + max([utilization, throughput, sum(response_time) / len(response_time), sum(waiting_time) / len(waiting_time), sum(turnaround_time) / len(turnaround_time), run_time]) * 0.01, f'{round(v, 2)}', color='blue', fontweight='bold')
    matplotlib.pyplot.bar([i for i in range(6)], [utilization, throughput, sum(response_time) / len(response_time), sum(waiting_time) / len(waiting_time), sum(turnaround_time) / len(turnaround_time), run_time], 0.5)
    mng = matplotlib.pyplot.get_current_fig_manager()
    mng.resize(*mng.window.maxsize())
    matplotlib.pyplot.show()

def SRT_process_data_chart(flag):
    data_for_chart = SRT_all(flag)
    completion_time = data_for_chart['completion_time']
    response_time = data_for_chart['response_time']
    waiting_time = data_for_chart['waiting_time']
    turnaround_time = data_for_chart['turnaround_time']
    matplotlib.pyplot.subplot(2, 2, 1) 
    matplotlib.pyplot.title('Completion Time') 
    matplotlib.pyplot.xlabel('Process Number') 
    matplotlib.pyplot.ylabel('Time') 
    matplotlib.pyplot.xticks([i for i in range(n)], [f'P{i}' for i in range(n)])
    matplotlib.pyplot.grid(True)
    for i, v in enumerate(completion_time):
        matplotlib.pyplot.text(i - 0.02, v + max(completion_time) * 0.01, f'{v}', color='blue', fontweight='bold')
    matplotlib.pyplot.bar([i for i in range(n)], completion_time, 0.5)
    matplotlib.pyplot.subplot(2, 2, 2) 
    matplotlib.pyplot.title('Response Time') 
    matplotlib.pyplot.xlabel('Process Number') 
    matplotlib.pyplot.ylabel('Time') 
    matplotlib.pyplot.xticks([i for i in range(n)], [f'P{i}' for i in range(n)])
    matplotlib.pyplot.grid(True)
    for i, v in enumerate(response_time):
        matplotlib.pyplot.text(i - 0.02, v + max(response_time) * 0.01, f'{v}', color='blue', fontweight='bold')
    matplotlib.pyplot.bar([i for i in range(n)], response_time, 0.5)
    matplotlib.pyplot.subplot(2, 2, 3) 
    matplotlib.pyplot.title('Waiting Time') 
    matplotlib.pyplot.xlabel('Process Number') 
    matplotlib.pyplot.ylabel('Time') 
    matplotlib.pyplot.xticks([i for i in range(n)], [f'P{i}' for i in range(n)])
    matplotlib.pyplot.grid(True)
    for i, v in enumerate(waiting_time):
        matplotlib.pyplot.text(i - 0.02, v + max(waiting_time) * 0.01, f'{v}', color='blue', fontweight='bold')
    matplotlib.pyplot.bar([i for i in range(n)], waiting_time, 0.5)
    matplotlib.pyplot.subplot(2, 2, 4) 
    matplotlib.pyplot.title('Turnaround Time') 
    matplotlib.pyplot.xlabel('Process Number') 
    matplotlib.pyplot.ylabel('Time') 
    matplotlib.pyplot.xticks([i for i in range(n)], [f'P{i}' for i in range(n)])
    matplotlib.pyplot.grid(True)
    for i, v in enumerate(turnaround_time):
        matplotlib.pyplot.text(i - 0.02, v + max(turnaround_time) * 0.01, f'{v}', color='blue', fontweight='bold')
    matplotlib.pyplot.bar([i for i in range(n)], turnaround_time, 0.5)
    mng = matplotlib.pyplot.get_current_fig_manager()
    mng.resize(*mng.window.maxsize())
    matplotlib.pyplot.show()

def SRT_algorithm_data_chart(flag):
    data_for_chart = SRT_all(flag)
    utilization = data_for_chart['utilization']
    throughput = data_for_chart['throughput']
    completion_time = data_for_chart['completion_time']
    response_time = data_for_chart['response_time']
    waiting_time = data_for_chart['waiting_time']
    turnaround_time = data_for_chart['turnaround_time']
    run_time = SRT_time(flag) * 1000
    matplotlib.pyplot.title('Algorithm Data') 
    matplotlib.pyplot.xlabel('Name') 
    matplotlib.pyplot.ylabel('Number') 
    matplotlib.pyplot.xticks([i for i in range(6)], ['Utilization', 'Throughput', 'Avg. Response Time', 'Avg. Waiting Time', 'Avg. Turnaround Time', 'Run Time * 1000'])
    matplotlib.pyplot.grid(True)
    for i, v in enumerate([utilization, throughput, sum(response_time) / len(response_time), sum(waiting_time) / len(waiting_time), sum(turnaround_time) / len(turnaround_time), run_time]):
        matplotlib.pyplot.text(i - 0.02, v + max([utilization, throughput, sum(response_time) / len(response_time), sum(waiting_time) / len(waiting_time), sum(turnaround_time) / len(turnaround_time), run_time]) * 0.01, f'{round(v, 2)}', color='blue', fontweight='bold')
    matplotlib.pyplot.bar([i for i in range(6)], [utilization, throughput, sum(response_time) / len(response_time), sum(waiting_time) / len(waiting_time), sum(turnaround_time) / len(turnaround_time), run_time], 0.5)
    mng = matplotlib.pyplot.get_current_fig_manager()
    mng.resize(*mng.window.maxsize())
    matplotlib.pyplot.show()

def RR_process_data_chart(time_quantum , flag):
    data_for_chart = RR_all(time_quantum , flag)
    completion_time = data_for_chart['completion_time']
    response_time = data_for_chart['response_time']
    waiting_time = data_for_chart['waiting_time']
    turnaround_time = data_for_chart['turnaround_time']
    matplotlib.pyplot.subplot(2, 2, 1) 
    matplotlib.pyplot.title('Completion Time') 
    matplotlib.pyplot.xlabel('Process Number') 
    matplotlib.pyplot.ylabel('Time') 
    matplotlib.pyplot.xticks([i for i in range(n)], [f'P{i}' for i in range(n)])
    matplotlib.pyplot.grid(True)
    for i, v in enumerate(completion_time):
        matplotlib.pyplot.text(i - 0.02, v + max(completion_time) * 0.01, f'{v}', color='blue', fontweight='bold')
    matplotlib.pyplot.bar([i for i in range(n)], completion_time, 0.5)
    matplotlib.pyplot.subplot(2, 2, 2) 
    matplotlib.pyplot.title('Response Time') 
    matplotlib.pyplot.xlabel('Process Number') 
    matplotlib.pyplot.ylabel('Time') 
    matplotlib.pyplot.xticks([i for i in range(n)], [f'P{i}' for i in range(n)])
    matplotlib.pyplot.grid(True)
    for i, v in enumerate(response_time):
        matplotlib.pyplot.text(i - 0.02, v + max(response_time) * 0.01, f'{v}', color='blue', fontweight='bold')
    matplotlib.pyplot.bar([i for i in range(n)], response_time, 0.5)
    matplotlib.pyplot.subplot(2, 2, 3) 
    matplotlib.pyplot.title('Waiting Time') 
    matplotlib.pyplot.xlabel('Process Number') 
    matplotlib.pyplot.ylabel('Time') 
    matplotlib.pyplot.xticks([i for i in range(n)], [f'P{i}' for i in range(n)])
    matplotlib.pyplot.grid(True)
    for i, v in enumerate(waiting_time):
        matplotlib.pyplot.text(i - 0.02, v + max(waiting_time) * 0.01, f'{v}', color='blue', fontweight='bold')
    matplotlib.pyplot.bar([i for i in range(n)], waiting_time, 0.5)
    matplotlib.pyplot.subplot(2, 2, 4) 
    matplotlib.pyplot.title('Turnaround Time') 
    matplotlib.pyplot.xlabel('Process Number') 
    matplotlib.pyplot.ylabel('Time') 
    matplotlib.pyplot.xticks([i for i in range(n)], [f'P{i}' for i in range(n)])
    matplotlib.pyplot.grid(True)
    for i, v in enumerate(turnaround_time):
        matplotlib.pyplot.text(i - 0.02, v + max(turnaround_time) * 0.01, f'{v}', color='blue', fontweight='bold')
    matplotlib.pyplot.bar([i for i in range(n)], turnaround_time, 0.5)
    mng = matplotlib.pyplot.get_current_fig_manager()
    mng.resize(*mng.window.maxsize())
    matplotlib.pyplot.show()

def RR_algorithm_data_chart(time_quantum , flag):
    data_for_chart = RR_all(time_quantum , flag)
    utilization = data_for_chart['utilization']
    throughput = data_for_chart['throughput']
    completion_time = data_for_chart['completion_time']
    response_time = data_for_chart['response_time']
    waiting_time = data_for_chart['waiting_time']
    turnaround_time = data_for_chart['turnaround_time']
    run_time = RR_time(time_quantum , flag) * 1000
    matplotlib.pyplot.title('Algorithm Data') 
    matplotlib.pyplot.xlabel('Name') 
    matplotlib.pyplot.ylabel('Number') 
    matplotlib.pyplot.xticks([i for i in range(6)], ['Utilization', 'Avg. Throughput', 'Avg. Response Time', 'Avg. Waiting Time', 'Avg. Turnaround Time', 'Run Time * 1000'])
    matplotlib.pyplot.grid(True)
    for i, v in enumerate([utilization, throughput, sum(response_time) / len(response_time), sum(waiting_time) / len(waiting_time), sum(turnaround_time) / len(turnaround_time), run_time]):
        matplotlib.pyplot.text(i - 0.02, v + max([utilization, throughput, sum(response_time) / len(response_time), sum(waiting_time) / len(waiting_time), sum(turnaround_time) / len(turnaround_time), run_time]) * 0.01, f'{round(v, 2)}', color='blue', fontweight='bold')
    matplotlib.pyplot.bar([i for i in range(6)], [utilization, throughput, sum(response_time) / len(response_time), sum(waiting_time) / len(waiting_time), sum(turnaround_time) / len(turnaround_time), run_time], 0.5)
    mng = matplotlib.pyplot.get_current_fig_manager()
    mng.resize(*mng.window.maxsize())
    matplotlib.pyplot.show()
    
def All_algorithms_algorithm_data_chart(time_quantum):
    dictionary = run_all(time_quantum)
    FCFS=dictionary['FCFS']
    SJF=dictionary['SJF']
    SRT=dictionary['SRT']
    RR=dictionary['RR']
    run_time = run_all_time(time_quantum ) * 1000
    
    FCFS_utilization = FCFS['utilization']
    FCFS_throughput = FCFS['throughput']
    FCFS_completion_time = FCFS['completion_time']
    FCFS_response_time = FCFS['response_time']
    FCFS_waiting_time = FCFS['waiting_time']
    FCFS_turnaround_time = FCFS['turnaround_time']
    
    SJF_utilization = SJF['utilization']
    SJF_throughput = SJF['throughput']
    SJF_completion_time = SJF['completion_time']
    SJF_response_time = SJF['response_time']
    SJF_waiting_time = SJF['waiting_time']
    SJF_turnaround_time = SJF['turnaround_time']
    
    SRT_utilization = SRT['utilization']
    SRT_throughput = SRT['throughput']
    SRT_completion_time = SRT['completion_time']
    SRT_response_time = SRT['response_time']
    SRT_waiting_time = SRT['waiting_time']
    SRT_turnaround_time = SRT['turnaround_time']
    
    RR_utilization = RR['utilization']
    RR_throughput = RR['throughput']
    RR_completion_time = RR['completion_time']
    RR_response_time = RR['response_time']
    RR_waiting_time = RR['waiting_time']
    RR_turnaround_time = RR['turnaround_time']
    
    matplotlib.pyplot.title('Algorithm Data') 
    matplotlib.pyplot.xlabel('Name') 
    matplotlib.pyplot.ylabel('Number') 
    matplotlib.pyplot.xticks([i for i in range(6)], ['Utilization', 'Avg. Throughput', 'Avg. Response Time', 'Avg. Waiting Time', 'Avg. Turnaround Time', 'Run Time * 1000'])
    matplotlib.pyplot.grid(True)
    
    for i, v in enumerate([FCFS_utilization, FCFS_throughput, sum(FCFS_response_time) / len(FCFS_response_time), sum(FCFS_waiting_time) / len(FCFS_waiting_time), sum(FCFS_turnaround_time) / len(FCFS_turnaround_time), run_time]):
        matplotlib.pyplot.text(i + 0.02, v + max([FCFS_utilization, FCFS_throughput, sum(FCFS_response_time) / len(FCFS_response_time), sum(FCFS_waiting_time) / len(FCFS_waiting_time), sum(FCFS_turnaround_time) / len(FCFS_turnaround_time), run_time]) * 0.01, f'{round(v, 2)}', color='blue', fontweight='bold')
    matplotlib.pyplot.bar([i+0.1 for i in range(6)], [FCFS_utilization, FCFS_throughput, sum(FCFS_response_time) / len(FCFS_response_time), sum(FCFS_waiting_time) / len(FCFS_waiting_time), sum(FCFS_turnaround_time) / len(FCFS_turnaround_time), run_time], 0.2 , label='FCFS')
    
    for i, v in enumerate([SJF_utilization, SJF_throughput, sum(SJF_response_time) / len(SJF_response_time), sum(SJF_waiting_time) / len(SJF_waiting_time), sum(SJF_turnaround_time) / len(SJF_turnaround_time), run_time]):
        matplotlib.pyplot.text(i - 0.2, v + max([SJF_utilization, SJF_throughput, sum(SJF_response_time) / len(SJF_response_time), sum(SJF_waiting_time) / len(SJF_waiting_time), sum(SJF_turnaround_time) / len(SJF_turnaround_time), run_time]) * 0.01, f'{round(v, 2)}', color='orange', fontweight='bold')
    matplotlib.pyplot.bar([i-0.1 for i in range(6)], [SJF_utilization, SJF_throughput, sum(SJF_response_time) / len(SJF_response_time), sum(SJF_waiting_time) / len(SJF_waiting_time), sum(SJF_turnaround_time) / len(SJF_turnaround_time), run_time], 0.2 , label='SJF')
    
    for i, v in enumerate([SRT_utilization, SRT_throughput, sum(SRT_response_time) / len(SRT_response_time), sum(SRT_waiting_time) / len(SRT_waiting_time), sum(SRT_turnaround_time) / len(SRT_turnaround_time), run_time]):
        matplotlib.pyplot.text(i - 0.4, v + max([SRT_utilization, SRT_throughput, sum(SRT_response_time) / len(SRT_response_time), sum(SRT_waiting_time) / len(SRT_waiting_time), sum(SRT_turnaround_time) / len(SRT_turnaround_time), run_time]) * 0.01, f'{round(v, 2)}', color='green', fontweight='bold')
    matplotlib.pyplot.bar([i-0.3 for i in range(6)], [SRT_utilization, SRT_throughput, sum(SRT_response_time) / len(SRT_response_time), sum(SRT_waiting_time) / len(SRT_waiting_time), sum(SRT_turnaround_time) / len(SRT_turnaround_time), run_time], 0.2 , label='SRT')
    
    for i, v in enumerate([RR_utilization, RR_throughput, sum(RR_response_time) / len(RR_response_time), sum(RR_waiting_time) / len(RR_waiting_time), sum(RR_turnaround_time) / len(RR_turnaround_time), run_time]):
        matplotlib.pyplot.text(i + 0.21, v + max([RR_utilization, RR_throughput, sum(RR_response_time) / len(RR_response_time), sum(RR_waiting_time) / len(RR_waiting_time), sum(RR_turnaround_time) / len(RR_turnaround_time), run_time]) * 0.01, f'{round(v, 2)}', color='red', fontweight='bold')
    matplotlib.pyplot.bar([i+0.3 for i in range(6)], [RR_utilization, RR_throughput, sum(RR_response_time) / len(RR_response_time), sum(RR_waiting_time) / len(RR_waiting_time), sum(RR_turnaround_time) / len(RR_turnaround_time), run_time], 0.2 , label='RR')
    
    mng = matplotlib.pyplot.get_current_fig_manager()
    matplotlib.pyplot.legend()
    mng.resize(*mng.window.maxsize())
    matplotlib.pyplot.show()
    return 
def All_algorithms_process_data_chart(time_quantum):
    dictionary = run_all(time_quantum)
    FCFS=dictionary['FCFS']
    SJF=dictionary['SJF']
    SRT=dictionary['SRT']
    RR=dictionary['RR']
    
    FCFS_completion_time = FCFS['completion_time']
    FCFS_response_time = FCFS['response_time']
    FCFS_waiting_time = FCFS['waiting_time']
    FCFS_turnaround_time = FCFS['turnaround_time']
    
    SJF_completion_time = SJF['completion_time']
    SJF_response_time = SJF['response_time']
    SJF_waiting_time = SJF['waiting_time']
    SJF_turnaround_time = SJF['turnaround_time']
    
    SRT_completion_time = SRT['completion_time']
    SRT_response_time = SRT['response_time']
    SRT_waiting_time = SRT['waiting_time']
    SRT_turnaround_time = SRT['turnaround_time']
    
    RR_completion_time = RR['completion_time']
    RR_response_time = RR['response_time']
    RR_waiting_time = RR['waiting_time']
    RR_turnaround_time = RR['turnaround_time']
    
    matplotlib.pyplot.subplot(2, 2, 1) 
    matplotlib.pyplot.title('Completion Time') 
    matplotlib.pyplot.xlabel('Process Number') 
    matplotlib.pyplot.ylabel('Time') 
    matplotlib.pyplot.xticks([i for i in range(n)], [f'P{i}' for i in range(n)])
    matplotlib.pyplot.grid(True)
    
    for i, v in enumerate(FCFS_completion_time):
        matplotlib.pyplot.text(i + 0.002, v + max(FCFS_completion_time) * 0.01, f'{v}', color='blue', fontweight='bold' )
    matplotlib.pyplot.bar([i+0.025 for i in range(n)], FCFS_completion_time, 0.05, label='FCFS')
    for i, v in enumerate(SJF_completion_time):
        matplotlib.pyplot.text(i - 0.055, v + max(SJF_completion_time) * 0.01, f'{v}', color='orange', fontweight='bold' )
    matplotlib.pyplot.bar([i-0.025 for i in range(n)], SJF_completion_time, 0.05, label='SJF')
    for i, v in enumerate(SRT_completion_time):
        matplotlib.pyplot.text(i + 0.052, v + max(SRT_completion_time) * 0.01, f'{v}', color='green', fontweight='bold' )
    matplotlib.pyplot.bar([i+.075 for i in range(n)], SRT_completion_time, 0.05, label='SRT')
    for i, v in enumerate(RR_completion_time):
        matplotlib.pyplot.text(i - 0.109, v + max(RR_completion_time) * 0.01, f'{v}', color='red', fontweight='bold' )
    matplotlib.pyplot.bar([i-0.075 for i in range(n)], RR_completion_time, 0.05, label='RR')
    matplotlib.pyplot.legend( shadow=True, prop={'size': 6})
    
    matplotlib.pyplot.subplot(2, 2, 2) 
    matplotlib.pyplot.title('Response Time') 
    matplotlib.pyplot.xlabel('Process Number') 
    matplotlib.pyplot.ylabel('Time') 
    matplotlib.pyplot.xticks([i for i in range(n)], [f'P{i}' for i in range(n)])
    matplotlib.pyplot.grid(True)
    
    for i, v in enumerate(FCFS_response_time):
        matplotlib.pyplot.text(i + 0.002, v + max(FCFS_response_time) * 0.01, f'{v}', color='blue', fontweight='bold' )
    matplotlib.pyplot.bar([i+0.025 for i in range(n)], FCFS_response_time, 0.05, label='FCFS')
    for i, v in enumerate(SJF_response_time):
        matplotlib.pyplot.text(i - 0.055, v + max(SJF_response_time) * 0.01, f'{v}', color='orange', fontweight='bold' )
    matplotlib.pyplot.bar([i-0.025 for i in range(n)], SJF_response_time, 0.05, label='SJF')
    for i, v in enumerate(SRT_response_time):
        matplotlib.pyplot.text(i + 0.052, v + max(SRT_response_time) * 0.01, f'{v}', color='green', fontweight='bold' )
    matplotlib.pyplot.bar([i+.075 for i in range(n)], SRT_response_time, 0.05, label='SRT')
    for i, v in enumerate(RR_response_time):
        matplotlib.pyplot.text(i - 0.109, v + max(RR_response_time) * 0.01, f'{v}', color='red', fontweight='bold' )
    matplotlib.pyplot.bar([i-0.075 for i in range(n)], RR_response_time, 0.05 , label='RR')
    matplotlib.pyplot.legend( shadow=True, prop={'size': 6})
    
    matplotlib.pyplot.subplot(2, 2, 3) 
    matplotlib.pyplot.title('Waiting Time') 
    matplotlib.pyplot.xlabel('Process Number') 
    matplotlib.pyplot.ylabel('Time') 
    matplotlib.pyplot.xticks([i for i in range(n)], [f'P{i}' for i in range(n)])
    matplotlib.pyplot.grid(True)
    
    for i, v in enumerate(FCFS_waiting_time):
        matplotlib.pyplot.text(i + 0.002, v + max(FCFS_waiting_time) * 0.01, f'{v}', color='blue', fontweight='bold' )
    matplotlib.pyplot.bar([i+0.025 for i in range(n)], FCFS_waiting_time, 0.05, label='FCFS')
    for i, v in enumerate(SJF_waiting_time):
        matplotlib.pyplot.text(i - 0.055, v + max(SJF_waiting_time) * 0.01, f'{v}', color='orange', fontweight='bold' )
    matplotlib.pyplot.bar([i-0.025 for i in range(n)], SJF_waiting_time, 0.05, label='SJF')
    for i, v in enumerate(SRT_waiting_time):
        matplotlib.pyplot.text(i + 0.052, v + max(SRT_waiting_time) * 0.01, f'{v}', color='green', fontweight='bold' )
    matplotlib.pyplot.bar([i+.075 for i in range(n)], SRT_waiting_time, 0.05, label='SRT')
    for i, v in enumerate(RR_waiting_time):
        matplotlib.pyplot.text(i - 0.109, v + max(RR_waiting_time) * 0.01, f'{v}', color='red', fontweight='bold' )
    matplotlib.pyplot.bar([i-0.075 for i in range(n)], RR_waiting_time, 0.05 , label='RR')
    matplotlib.pyplot.legend( shadow=True, prop={'size': 6})
    
    matplotlib.pyplot.subplot(2, 2, 4) 
    matplotlib.pyplot.title('Turn Around Time') 
    matplotlib.pyplot.xlabel('Process Number') 
    matplotlib.pyplot.ylabel('Time') 
    matplotlib.pyplot.xticks([i for i in range(n)], [f'P{i}' for i in range(n)])
    matplotlib.pyplot.grid(True)
    
    for i, v in enumerate(FCFS_turnaround_time):
        matplotlib.pyplot.text(i + 0.002, v + max(FCFS_turnaround_time) * 0.01, f'{v}', color='blue', fontweight='bold' )
    matplotlib.pyplot.bar([i+0.025 for i in range(n)], FCFS_turnaround_time, 0.05, label='FCFS')
    for i, v in enumerate(SJF_turnaround_time):
        matplotlib.pyplot.text(i - 0.055, v + max(SJF_turnaround_time) * 0.01, f'{v}', color='orange', fontweight='bold' )
    matplotlib.pyplot.bar([i-0.025 for i in range(n)], SJF_turnaround_time, 0.05, label='SJF')
    for i, v in enumerate(SRT_turnaround_time):
        matplotlib.pyplot.text(i + 0.052, v + max(SRT_turnaround_time) * 0.01, f'{v}', color='green', fontweight='bold' )
    matplotlib.pyplot.bar([i+.075 for i in range(n)], SRT_turnaround_time, 0.05, label='SRT')
    for i, v in enumerate(RR_turnaround_time):
        matplotlib.pyplot.text(i - 0.109, v + max(RR_turnaround_time) * 0.01, f'{v}', color='red', fontweight='bold' )
    matplotlib.pyplot.bar([i-0.075 for i in range(n)], RR_turnaround_time, 0.05 , label='RR')
    matplotlib.pyplot.legend( shadow=True, prop={'size': 6})
    
        
    mng = matplotlib.pyplot.get_current_fig_manager()
    mng.resize(*mng.window.maxsize())
    
    matplotlib.pyplot.show()
    return

def FCFS_commands():
    FCFS_screen = tkinter.Tk()
    FCFS_screen.configure(background='#222f3e')
    FCFS_screen.attributes("-fullscreen", True)
    
    FCFS_label = tkinter.Label(FCFS_screen, text='First Come First Served', bg='#222f3e', fg='#8395a7', highlightthickness=0, font="-weight bold")
    FCFS_label.pack(ipadx=68, ipady=20, pady=(120, 30))

    FCFS_gantt_button = tkinter.Button(FCFS_screen, command=draw_FCFS_chart, text='Draw Gantt Chart', bg='#576574', fg='#8395a7', highlightthickness=0, font="-weight bold")
    FCFS_gantt_button.pack(ipadx=68, ipady=20, pady=20)

    FCFS_process_data_button = tkinter.Button(FCFS_screen, command=lambda :FCFS_process_data_chart(0), text='Draw Process Data Chart', bg='#576574', fg='#8395a7', highlightthickness=0, font="-weight bold")
    FCFS_process_data_button.pack(ipadx=35, ipady=20, pady=20)

    FCFS_algorithm_data_button = tkinter.Button(FCFS_screen, command=lambda:FCFS_algorithm_data_chart(0), text='Draw Algorithm Data Chart', bg='#576574', fg='#8395a7', highlightthickness=0, font="-weight bold")
    FCFS_algorithm_data_button.pack(ipadx=25, ipady=20, pady=20)

    FCFS_exit_button = tkinter.Button(FCFS_screen, text='BACK', command=FCFS_screen.destroy, bg='#576574',fg='#8395a7', highlightthickness=0, font="-weight bold")
    FCFS_exit_button.pack(ipadx=115, ipady=20, pady=20)

    return

def SJF_commands():
    SJF_screen = tkinter.Tk()
    SJF_screen.configure(background='#222f3e')
    SJF_screen.attributes("-fullscreen", True)

    SJF_label = tkinter.Label(SJF_screen, text='Shortest Job First', bg='#222f3e', fg='#8395a7', highlightthickness=0, font="-weight bold")
    SJF_label.pack(ipadx=68, ipady=20, pady=(120, 30))
    
    SJF_gantt_button = tkinter.Button(SJF_screen, command=draw_SJF_chart, text='Draw Gantt Chart', bg='#576574', fg='#8395a7', highlightthickness=0, font="-weight bold")
    SJF_gantt_button.pack(ipadx=68, ipady=20, pady=20)

    SJF_process_data_button = tkinter.Button(SJF_screen, command=lambda : SJF_process_data_chart(0), text='Draw Process Data Chart', bg='#576574', fg='#8395a7', highlightthickness=0, font="-weight bold")
    SJF_process_data_button.pack(ipadx=35, ipady=20, pady=20)

    SJF_algorithm_data_button = tkinter.Button(SJF_screen, command= lambda:SJF_algorithm_data_chart(0), text='Draw Algorithm Data Chart', bg='#576574', fg='#8395a7', highlightthickness=0, font="-weight bold")
    SJF_algorithm_data_button.pack(ipadx=25, ipady=20, pady=20)

    SJF_exit_button = tkinter.Button(SJF_screen, text='BACK', command=SJF_screen.destroy, bg='#576574',fg='#8395a7', highlightthickness=0, font="-weight bold")
    SJF_exit_button.pack(ipadx=115, ipady=20, pady=20)

    return

def SRT_commands():
    SRT_screen = tkinter.Tk()
    SRT_screen.configure(background='#222f3e')
    SRT_screen.attributes("-fullscreen", True)

    SRT_label = tkinter.Label(SRT_screen, text='Shortest Remainig Time', bg='#222f3e', fg='#8395a7', highlightthickness=0, font="-weight bold")
    SRT_label.pack(ipadx=68, ipady=20, pady=(120, 30))
    
    SRT_gantt_button = tkinter.Button(SRT_screen, command=draw_SRT_chart, text='Draw Gantt Chart', bg='#576574', fg='#8395a7', highlightthickness=0, font="-weight bold")
    SRT_gantt_button.pack(ipadx=68, ipady=20, pady=20)

    SRT_process_data_button = tkinter.Button(SRT_screen, command=lambda :SRT_process_data_chart(0), text='Draw Process Data Chart', bg='#576574', fg='#8395a7', highlightthickness=0, font="-weight bold")
    SRT_process_data_button.pack(ipadx=35, ipady=20, pady=20)

    SRT_algorithm_data_button = tkinter.Button(SRT_screen, command=lambda :SRT_algorithm_data_chart(0), text='Draw Algorithm Data Chart', bg='#576574', fg='#8395a7', highlightthickness=0, font="-weight bold")
    SRT_algorithm_data_button.pack(ipadx=25, ipady=20, pady=20)

    SRT_exit_button = tkinter.Button(SRT_screen, text='BACK', command=SRT_screen.destroy, bg='#576574',fg='#8395a7', highlightthickness=0, font="-weight bold")
    SRT_exit_button.pack(ipadx=115, ipady=20, pady=20)

    return

def RR_commands():
    RR_screen = tkinter.Tk()
    RR_screen.configure(background='#222f3e')
    RR_screen.attributes("-fullscreen", True)

    RR_label = tkinter.Label(RR_screen, text='Round Robbin', bg='#222f3e', fg='#8395a7', highlightthickness=0, font="-weight bold")
    RR_label.pack(ipadx=68, ipady=20, pady=(120, 30))

    time_quantum_entry = tkinter.Entry(RR_screen, width=31, highlightthickness=0, font="-weight bold")
    time_quantum_entry.insert(0, ' -->  Enter Time Quantum Here  <-- ')
    time_quantum_entry.select_clear()
    time_quantum_entry.pack(ipady=20, pady=20)
    
    RR_gantt_button = tkinter.Button(RR_screen, command=lambda: draw_RR_chart(int(time_quantum_entry.get())), text='Draw Gantt Chart', bg='#576574', fg='#8395a7', highlightthickness=0, font="-weight bold")
    RR_gantt_button.pack(ipadx=68, ipady=20, pady=20)

    RR_process_data_button = tkinter.Button(RR_screen, command=lambda: RR_process_data_chart(int(time_quantum_entry.get()) ,0), text='Draw Process Data Chart', bg='#576574', fg='#8395a7', highlightthickness=0, font="-weight bold")
    RR_process_data_button.pack(ipadx=35, ipady=20, pady=20)

    RR_algorithm_data_button = tkinter.Button(RR_screen, command=lambda: RR_algorithm_data_chart(int(time_quantum_entry.get()) , 0), text='Draw Algorithm Data Chart', bg='#576574', fg='#8395a7', highlightthickness=0, font="-weight bold")
    RR_algorithm_data_button.pack(ipadx=25, ipady=20, pady=20)

    RR_exit_button = tkinter.Button(RR_screen, text='BACK', command=RR_screen.destroy, bg='#576574',fg='#8395a7', highlightthickness=0, font="-weight bold")
    RR_exit_button.pack(ipadx=115, ipady=20, pady=20)

    return

def All_algorithms_command():
    All_algorithms_screen = tkinter.Tk()
    All_algorithms_screen.configure(background='#222f3e')
    All_algorithms_screen.attributes("-fullscreen", True)
    
    All_algorithms_label = tkinter.Label(All_algorithms_screen, text='All Algorithms', bg='#222f3e', fg='#8395a7', highlightthickness=0, font="-weight bold")
    All_algorithms_label.pack(ipadx=68, ipady=20, pady=(120, 30))
    
    time_quantum_entry = tkinter.Entry(All_algorithms_screen, width=31, highlightthickness=0, font="-weight bold")
    time_quantum_entry.insert(0, ' -->  Enter Time Quantum Here For RR <-- ')
    time_quantum_entry.select_clear()
    time_quantum_entry.pack(ipady=20, pady=20)

    All_algorithms_gantt_button = tkinter.Button(All_algorithms_screen, command= lambda :draw_All_algorithms_chart(int(time_quantum_entry.get())), text='Draw Gantt Chart', bg='#576574', fg='#8395a7', highlightthickness=0, font="-weight bold")
    All_algorithms_gantt_button.pack(ipadx=68, ipady=20, pady=20)

    All_algorithms_process_data_button = tkinter.Button(All_algorithms_screen, command= lambda: All_algorithms_process_data_chart(int(time_quantum_entry.get())), text='Draw Process Data Chart', bg='#576574', fg='#8395a7', highlightthickness=0, font="-weight bold")
    All_algorithms_process_data_button.pack(ipadx=35, ipady=20, pady=20)

    All_algorithms_algorithm_data_button = tkinter.Button(All_algorithms_screen, command= lambda: All_algorithms_algorithm_data_chart(int(time_quantum_entry.get()) ), text='Draw Algorithm Data Chart', bg='#576574', fg='#8395a7', highlightthickness=0, font="-weight bold")
    All_algorithms_algorithm_data_button.pack(ipadx=25, ipady=20, pady=20)

    All_algorithms_exit_button = tkinter.Button(All_algorithms_screen, text='BACK', command=All_algorithms_screen.destroy, bg='#576574',fg='#8395a7', highlightthickness=0, font="-weight bold")
    All_algorithms_exit_button.pack(ipadx=115, ipady=20, pady=20)
    
    return

root = tkinter.Tk()
root.configure(background='#222f3e')
root.attributes("-fullscreen", True)

def home_page():
    home_page = tkinter.Tk()
    home_page.configure(background='#222f3e')
    home_page.attributes("-fullscreen", True)

    home_label = tkinter.Label(home_page, text='CPU Scheduling Algorithms Simulator | Home Page', bg='#222f3e', fg='#8395a7', highlightthickness=0, font="-weight bold")
    home_label.pack(ipadx=68, ipady=20, pady=(120, 30))

    FCFS_button = tkinter.Button(home_page, text='FCFS Algorithm', command=FCFS_commands, bg='#576574', fg='#8395a7', highlightthickness=0, font="-weight bold")
    FCFS_button.pack(ipadx=68, ipady=20, pady=20)

    SJF_button = tkinter.Button(home_page, text='SJF Algorithm', command=SJF_commands, bg='#576574',fg='#8395a7', highlightthickness=0, font="-weight bold")
    SJF_button.pack(ipadx=74, ipady=20, pady=20)

    SRT_button = tkinter.Button(home_page, text='SRT Algorithm', command=SRT_commands, bg='#576574',fg='#8395a7', highlightthickness=0, font="-weight bold")
    SRT_button.pack(ipadx=73, ipady=20, pady=20)

    RR_button = tkinter.Button(home_page, text='RR Algorithm', bg='#576574', command=RR_commands, fg='#8395a7', highlightthickness=0, font="-weight bold")
    RR_button.pack(ipadx=77, ipady=20, pady=20)

    all_button = tkinter.Button(home_page, text='All Algorithms', bg='#576574', command =All_algorithms_command ,fg='#8395a7', highlightthickness=0, font="-weight bold")
    all_button.pack(ipadx=74, ipady=20, pady=20)

    back_button = tkinter.Button(home_page, text='BACK', command=home_page.destroy, bg='#576574',fg='#8395a7', highlightthickness=0, font="-weight bold")
    back_button.pack(ipadx=115, ipady=20, pady=20)


def reader_helper():
    root.filename = filedialog.askopenfile(initialdir='.', title='Select an xlsx file')
    temp = str(root.filename).split("/")[-1]
    temp = temp.split('\'')[0]
    read_from_file(temp)
    home_page()
    return

def end():
    root.destroy()
    root.quit()

root_label = tkinter.Label(root, text='Welcome to CPU Scheduling Algorithms Simulator', bg='#222f3e', fg='#8395a7', highlightthickness=0, font="-weight bold")
root_label.pack(ipadx=68, ipady=20, pady=(120, 30))

read_button = tkinter.Button(root, command=reader_helper, text='Select an Excel File', bg='#576574', fg='#8395a7', highlightthickness=0, font="-weight bold")
read_button.pack(ipadx=52, ipady=20, pady=20)

exit_button = tkinter.Button(root, text='EXIT', command=end, bg='#576574',fg='#8395a7', highlightthickness=0, font="-weight bold")
exit_button.pack(ipadx=115, ipady=20, pady=20)

root.mainloop()